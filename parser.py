"""Lectura y consolidacion del pedido (hoja 'Picking Subcedis W##')."""
import re
import pandas as pd
import openpyxl

SHEET_PATTERN = re.compile(r"picking\s*subcedis\s*w\d+", re.IGNORECASE)

EXPECTED_COLS = [
    "id_cabecera",
    "id_linea",
    "codigo_departamento",
    "nombre_departamento",
    "codigo_color",
    "unidades_solicitadas",
    "unidades_recibidas",
    "articulo_original",
    "cabecera_original",
]


def find_picking_sheet(xlsx_path_or_buffer):
    """Devuelve el nombre de hoja que matchea 'Picking Subcedis W##' (variable)."""
    wb = openpyxl.load_workbook(xlsx_path_or_buffer, data_only=True, read_only=True)
    for name in wb.sheetnames:
        if SHEET_PATTERN.search(name):
            return name
    raise ValueError(
        "No se encontro ninguna hoja tipo 'Picking Subcedis W##' en el archivo. "
        f"Hojas disponibles: {wb.sheetnames}"
    )


def extract_week_tag(sheet_name):
    m = re.search(r"w\d+", sheet_name, re.IGNORECASE)
    return m.group(0).upper() if m else sheet_name


def quitar_punto(codigo):
    if codigo is None:
        return ""
    return str(codigo).replace(".", "").strip()


def _partir_codigo(codigo_color):
    """Separa 'cod' y 'color' a partir de codigo_color (ej. '146590.056' -> ('146590','056')).
    Si no hay punto, 'color' queda vacío."""
    if codigo_color is None:
        return "", ""
    texto = str(codigo_color).strip()
    if "." in texto:
        cod, color = texto.split(".", 1)
        return cod, color
    return texto, ""


def cargar_y_consolidar(xlsx_path_or_buffer):
    """Lee el pedido y devuelve (week_tag, df_consolidado, df_detalle_crudo).

    df_consolidado columns: tienda, nombre_tienda, codigo, cantidad_solicitada
    Agrupado por tienda + codigo (sin punto), sumando unidades_solicitadas.
    Se usa para la validación y el CSV del WMS.

    df_detalle_crudo: una fila por cada línea original del pedido (sin
    consolidar, tal como venía en el Excel), con columnas id_cabecera,
    id_linea, codigo_departamento, nombre_departamento, codigo_color,
    codigo (sin punto), unidades_solicitadas, unidades_recibidas,
    cabecera_original, articulo_original, cod, color.
    Se usa solo para el reporte descargable (pestaña de detalle).
    """
    sheet_name = find_picking_sheet(xlsx_path_or_buffer)
    week_tag = extract_week_tag(sheet_name)

    df = pd.read_excel(xlsx_path_or_buffer, sheet_name=sheet_name, dtype=str)

    missing = [c for c in EXPECTED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas esperadas en la hoja: {missing}")

    df["unidades_solicitadas"] = pd.to_numeric(df["unidades_solicitadas"], errors="coerce").fillna(0)
    df["tienda"] = df["codigo_departamento"].astype(str).str.strip()
    df["nombre_tienda"] = df["nombre_departamento"].astype(str).str.strip()
    df["codigo"] = df["codigo_color"].apply(quitar_punto)

    df = df[df["codigo"] != ""]

    consolidado = (
        df.groupby(["tienda", "nombre_tienda", "codigo"], as_index=False)["unidades_solicitadas"]
        .sum()
        .rename(columns={"unidades_solicitadas": "cantidad_solicitada"})
    )
    consolidado = consolidado.sort_values(["tienda", "codigo"]).reset_index(drop=True)

    partidos = df["codigo_color"].apply(_partir_codigo)
    detalle = pd.DataFrame(
        {
            "id_cabecera": df["id_cabecera"],
            "id_linea": df["id_linea"],
            "codigo_departamento": df["tienda"],
            "nombre_departamento": df["nombre_tienda"],
            "codigo_color": df["codigo_color"],
            "codigo": df["codigo"],
            "unidades_solicitadas": df["unidades_solicitadas"],
            "unidades_recibidas": pd.to_numeric(df["unidades_recibidas"], errors="coerce").fillna(0),
            "cabecera_original": df["cabecera_original"],
            "articulo_original": df["articulo_original"],
            "cod": [p[0] for p in partidos],
            "color": [p[1] for p in partidos],
        }
    ).reset_index(drop=True)

    return week_tag, consolidado, detalle


def generar_csv_wms(consolidado_df, fecha_emision, fecha_entrega):
    """Genera el DataFrame en formato plantilla WMS (PDE_*).

    fecha_emision / fecha_entrega: objetos date.
    PDE_num_doc = tienda + '-' + ddMM(fecha_emision)
    PDE_lin_doc = 1 (fijo)
    PDE_cod_tdo_rel = 'PD' (fijo)
    """
    ddmm = fecha_emision.strftime("%d%m")
    fec_emi_str = fecha_emision.strftime("%d/%m/%Y")
    fec_ent_str = fecha_entrega.strftime("%d/%m/%Y")

    rows = []
    for _, r in consolidado_df.iterrows():
        rows.append(
            {
                "PDE_num_doc": f"{r['tienda']}-{ddmm}",
                "PDE_lin_doc": 1,
                "PDE_fec_emi": fec_emi_str,
                "PDE_fec_ent": fec_ent_str,
                "PDE_cod_mat": r["codigo"],
                "PDE_pdt_mat": int(r["cantidad_solicitada"]) if float(r["cantidad_solicitada"]).is_integer() else r["cantidad_solicitada"],
                "PDE_cod_tdo_rel": "PD",
            }
        )
    return pd.DataFrame(rows, columns=[
        "PDE_num_doc", "PDE_lin_doc", "PDE_fec_emi", "PDE_fec_ent",
        "PDE_cod_mat", "PDE_pdt_mat", "PDE_cod_tdo_rel",
    ])
