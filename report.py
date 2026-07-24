"""
Genera el reporte descargable (Excel) de una semana: RESUMEN por tienda
(solicitado + resultados de validación) y DETALLE del pedido consolidado.
"""
import io

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _ultimo_cierre_por_tienda(historial_rows):
    """historial_rows: lista de tuplas
    (week_tag, tienda, fecha_cierre, solicitado_total, tenido_total, faltante_total, devuelto_total)
    Ya vienen ordenadas por fecha_cierre descendente (más reciente primero),
    así que basta quedarnos con la primera aparición de cada tienda."""
    ultimo = {}
    for row in historial_rows:
        _, tienda, _fecha, solicitado, tenido, falta, devuelto = row
        if tienda not in ultimo:
            ultimo[tienda] = {
                "solicitado_val": solicitado,
                "tenido": tenido,
                "falta": falta,
                "devuelto": devuelto,
            }
    return ultimo


def generar_reporte(db, conn, week_tag):
    """Devuelve (bytes_xlsx, resumen_df, detalle_df) para la semana dada.

    El detalle solo incluye los códigos que efectivamente se enviaron según
    la última validación cerrada de cada tienda (tenido > 0), y la cantidad
    mostrada es la enviada (tenido), no la solicitada originalmente. Las
    tiendas sin validación cerrada no aparecen en el detalle (nada
    confirmado como enviado todavía)."""
    tiendas = db.list_tiendas(conn, week_tag)
    historial_rows = db.get_historial(conn, week_tag=week_tag)
    ultimo_cierre = _ultimo_cierre_por_tienda(historial_rows)

    resumen_rows = []

    # Pedido crudo (una fila por línea original) para sacar las columnas
    # descriptivas (codigo_color, cod, color, id_cabecera, id_linea, etc.).
    # Nos quedamos con una fila por código+tienda (la primera), ya que la
    # cantidad real a mostrar viene de la validación, no de estas líneas.
    detalle_crudo = db.get_pedido_detalle(conn, week_tag)
    if not detalle_crudo.empty:
        detalle_crudo = detalle_crudo.drop_duplicates(
            subset=["codigo_departamento", "codigo"], keep="first"
        )

    detalle_rows = []

    # Una sola lectura del historial completo de la semana (con detalle por
    # código), en vez de una lectura por tienda — evita agotar la cuota de
    # la API de Google Sheets cuando hay muchas tiendas.
    detalle_validacion_por_tienda = db.get_ultimo_detalle_validacion_todas(conn, week_tag)

    for tienda, nombre in tiendas:
        pedido_map = db.get_pedido_tienda(conn, week_tag, tienda)
        cuenta_codigos = len(pedido_map)
        suma_solicitada = sum(pedido_map.values())

        cierre = ultimo_cierre.get(tienda)
        resumen_rows.append(
            {
                "codigo_departamento": tienda,
                "nombre_departamento": nombre,
                "Cuenta de codigo_color": cuenta_codigos,
                "Suma de unidades_solicitadas": suma_solicitada,
                "Tenido (validado)": cierre["tenido"] if cierre else None,
                "Falta": cierre["falta"] if cierre else None,
                "Devuelto": cierre["devuelto"] if cierre else None,
                "Validación cerrada": "Sí" if cierre else "No",
            }
        )

        # Detalle de códigos efectivamente enviados (tenido > 0) en la
        # última validación cerrada de esta tienda.
        detalle_validacion = detalle_validacion_por_tienda.get(str(tienda))
        if not detalle_validacion:
            continue

        for item in detalle_validacion:
            tenido = item.get("tenido", 0) or 0
            if tenido <= 0:
                continue
            codigo = item.get("codigo")
            fila_base = {}
            if not detalle_crudo.empty:
                match = detalle_crudo[
                    (detalle_crudo["codigo_departamento"] == tienda) & (detalle_crudo["codigo"] == codigo)
                ]
                if not match.empty:
                    fila_base = match.iloc[0].to_dict()

            detalle_rows.append(
                {
                    "id_cabecera": fila_base.get("id_cabecera", ""),
                    "id_linea": fila_base.get("id_linea", ""),
                    "codigo_departamento": tienda,
                    "nombre_departamento": nombre,
                    "codigo_color": fila_base.get("codigo_color", codigo),
                    "codigo": codigo,
                    "unidades_picking": tenido,
                    "cabecera_original": fila_base.get("cabecera_original", ""),
                    "articulo_original": fila_base.get("articulo_original", ""),
                    "cod": fila_base.get("cod", ""),
                    "color": fila_base.get("color", ""),
                }
            )

    resumen_df = pd.DataFrame(resumen_rows)
    detalle_df = pd.DataFrame(
        detalle_rows,
        columns=[
            "id_cabecera", "id_linea", "codigo_departamento", "nombre_departamento",
            "codigo_color", "codigo", "unidades_picking",
            "cabecera_original", "articulo_original", "cod", "color",
        ],
    )

    # Fila de totales generales
    if not resumen_df.empty:
        totales = {
            "codigo_departamento": "Total general",
            "nombre_departamento": "",
            "Cuenta de codigo_color": resumen_df["Cuenta de codigo_color"].sum(),
            "Suma de unidades_solicitadas": resumen_df["Suma de unidades_solicitadas"].sum(),
            "Tenido (validado)": resumen_df["Tenido (validado)"].sum(skipna=True),
            "Falta": resumen_df["Falta"].sum(skipna=True),
            "Devuelto": resumen_df["Devuelto"].sum(skipna=True),
            "Validación cerrada": "",
        }
        resumen_df = pd.concat([resumen_df, pd.DataFrame([totales])], ignore_index=True)

    # ------------------------------------------------------------------
    # Escribir el Excel con formato simple (encabezados en negrita/color)
    # ------------------------------------------------------------------
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        resumen_df.to_excel(writer, sheet_name="RESUMEN", index=False)
        detalle_df.to_excel(writer, sheet_name=f"Picking Subcedis {week_tag}", index=False)

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(
                (resumen_df.columns if sheet_name == "RESUMEN" else detalle_df.columns), start=1
            ):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                # ancho de columna aproximado según el contenido
                width = max(12, len(str(col_name)) + 2)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    return buffer.getvalue(), resumen_df, detalle_df
