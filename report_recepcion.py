"""
Genera el reporte descargable (Excel) de un documento de RECEPCIÓN:
 - DETALLE: lo que se escaneó tal cual, código por código y caja por caja,
   así haya diferencias con el packing list.
 - RESUMEN: comparación a nivel de código (sin importar la caja) entre lo
   esperado según el packing list y lo escaneado físicamente.
"""
import io

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def generar_reporte_recepcion(db, conn, documento):
    """Devuelve (bytes_xlsx, resumen_df, detalle_df, stock_warning) para el
    documento dado."""
    detalle_rows = db.get_detalle_documento_todas_cajas(conn, documento)
    scans_rows = db.get_scans_documento_todas_cajas(conn, documento)

    descripciones = db.get_stock_descripciones() if hasattr(db, "get_stock_descripciones") else {}
    stock_warning = db.get_stock_error() if hasattr(db, "get_stock_error") else None

    # ------------------------------------------------------------------
    # DETALLE: documento, codigo, descripcion, caja, unidades — tal cual se
    # escaneó (recibido + excedente), sin importar si coincide o no con lo
    # que decía el packing list.
    # ------------------------------------------------------------------
    detalle_rows_out = []
    for s in scans_rows:
        total_unidades = (s.get("recibido") or 0) + (s.get("devuelto") or 0)
        if total_unidades <= 0:
            continue
        codigo = s["codigo"]
        detalle_rows_out.append({
            "documento": documento,
            "codigo": codigo,
            "descripcion": descripciones.get(codigo, ""),
            "caja": s["box_number"],
            "unidades": int(total_unidades) if float(total_unidades).is_integer() else total_unidades,
        })

    detalle_df = pd.DataFrame(
        detalle_rows_out, columns=["documento", "codigo", "descripcion", "caja", "unidades"]
    )
    if not detalle_df.empty:
        detalle_df = detalle_df.sort_values(["caja", "codigo"]).reset_index(drop=True)

    # ------------------------------------------------------------------
    # RESUMEN: por código (agrupando todas las cajas), esperado según el
    # packing list vs escaneado físico. Para códigos que solo vienen dentro
    # de un "pool" (grupo de códigos con un tope compartido por caja, sin
    # cantidad individual definida en el packing list), no hay un esperado
    # por código — se marcan con tipo "pool" y el esperado queda vacío.
    # ------------------------------------------------------------------
    esperado_individual = {}
    tipo_por_codigo = {}
    for d in detalle_rows:
        codigo = d["codigo"]
        if d["tipo_linea"] == "individual":
            esperado_individual[codigo] = esperado_individual.get(codigo, 0) + (d.get("cantidad_esperada") or 0)
            tipo_por_codigo[codigo] = "individual"
        else:
            tipo_por_codigo.setdefault(codigo, "pool")

    escaneado_total = {}
    for s in scans_rows:
        codigo = s["codigo"]
        escaneado_total[codigo] = escaneado_total.get(codigo, 0) + (s.get("recibido") or 0) + (s.get("devuelto") or 0)

    todos_codigos = sorted(set(tipo_por_codigo) | set(escaneado_total))

    resumen_rows = []
    for codigo in todos_codigos:
        tipo = tipo_por_codigo.get(codigo, "pool")
        esperado = esperado_individual.get(codigo) if tipo == "individual" else None
        escaneado = escaneado_total.get(codigo, 0)
        diferencia = (escaneado - esperado) if esperado is not None else None
        resumen_rows.append({
            "codigo": codigo,
            "descripcion": descripciones.get(codigo, ""),
            "tipo": tipo,
            "esperado (packing list)": esperado,
            "escaneado (físico)": escaneado,
            "diferencia": diferencia,
        })

    resumen_df = pd.DataFrame(resumen_rows)

    if not resumen_df.empty:
        totales = {
            "codigo": "Total general",
            "descripcion": "",
            "tipo": "",
            "esperado (packing list)": resumen_df["esperado (packing list)"].sum(skipna=True),
            "escaneado (físico)": resumen_df["escaneado (físico)"].sum(skipna=True),
            "diferencia": resumen_df["diferencia"].sum(skipna=True),
        }
        resumen_df = pd.concat([resumen_df, pd.DataFrame([totales])], ignore_index=True)

    # ------------------------------------------------------------------
    # Escribir el Excel (mismo formato simple que el reporte de Picking)
    # ------------------------------------------------------------------
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        resumen_df.to_excel(writer, sheet_name="RESUMEN", index=False)
        detalle_df.to_excel(writer, sheet_name="DETALLE", index=False)

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)

        columnas_por_hoja = {"RESUMEN": resumen_df.columns, "DETALLE": detalle_df.columns}

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(columnas_por_hoja[sheet_name], start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                width = max(12, len(str(col_name)) + 2)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    return buffer.getvalue(), resumen_df, detalle_df, stock_warning
